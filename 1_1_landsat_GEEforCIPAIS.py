# ===============================================================
# Landsat 8–9 Land Surface Temperature (LST) processing pipeline
# Collection 2 Level-2 (L2SP) — Google Earth Engine
# Barsi atmospheric correction + in-situ calibration (SIMILE protocol)
# CIPAIS report — Lake Lugano
# Version of 16.12.2025, created by Yara Leone
# ===============================================================

import ee, geemap, os, rasterio, geopandas as gpd, numpy as np, time, shutil
from rasterio.mask import mask
from shapely.ops import unary_union
from datetime import datetime
import pandas as pd
import requests, tempfile
from openpyxl import load_workbook   

# -----------------------------
# GLOBAL PARAMETERS
# -----------------------------
LAKE_SHP   = r"inputs\SIMILE LAGO LUGANO SHP\SIMILE_laghi_UTM32.shp"
os.makedirs("outputs", exist_ok=True) 
OUT_BASE   = r"outputs\landsat8-9_L2SP"

YEAR = 2024
CLOUD_MAX = 80
BANK_ERODE_M = 90
PAUSE_S = 10

USE_ST_B10_FOR_FINAL_LST = False   # "True" if the Barsi method does not work

# Barsi constants (atmospheric parameters now taken from ST_ATRAN, ST_DRAD, ST_URAD)
EMIS = 0.98            # emissivity of water
K1   = 1321.08         # Barsi constant
K2   = 774.89          # Barsi constant

# Atmospheric scale factors (Landsat C2 L2 standard for 2024)
ATRAN_SCALE = 0.0001
DRAD_SCALE  = 0.001
URAD_SCALE  = 0.001

os.makedirs(OUT_BASE, exist_ok=True)

# -------------------------------------------------------------------
# Annual calibration (satellite → in-situ regression)
# Excel columns used for graphs: AB (estimated) and AC (in-situ)
# -------------------------------------------------------------------
CALIB_EXCEL = rf"inputs\temp_observed_estimated_{YEAR}.xlsx"
CALIB_X_COL = "AC"
CALIB_Y_COL = "AB"

# -----------------------------
# INITIALIZE EARTH ENGINE
# -----------------------------
try:
    ee.Initialize()
except Exception:
    ee.Authenticate(); ee.Initialize()

# -----------------------------
# LOAD LAKE SHAPEFILE
# -----------------------------
print("Loading Lugano lake shapefile…")
gdf = gpd.read_file(LAKE_SHP)

if gdf.crs is None:
    raise SystemExit("Shapefile has no CRS. Please define it (likely EPSG:32632).")

gdf_ee = gdf.to_crs(4326)
gdf_local = gdf

lake_union_ee = unary_union(list(gdf_ee.geometry))

for erosion_m in [BANK_ERODE_M, 30, 0]:
    erosion_deg = erosion_m / 111000.0
    lake_eroded_ee = lake_union_ee.buffer(-erosion_deg) if erosion_m > 0 else lake_union_ee
    if not lake_eroded_ee.is_empty:
        break

def shapely_to_ee(geom):
    if geom.geom_type == "Polygon":
        return ee.Geometry.Polygon(np.array(geom.exterior.coords).tolist())
    elif geom.geom_type == "MultiPolygon":
        polys=[]
        for p in geom.geoms:
            polys.append(ee.Geometry.Polygon(np.array(p.exterior.coords).tolist()))
        return ee.Geometry.MultiPolygon([p.coordinates().getInfo()[0] for p in polys])
    else:
        raise ValueError("Unsupported geometry type.")

lake_geom_ee = shapely_to_ee(lake_eroded_ee)
region = lake_geom_ee.bounds()

# -----------------------------------------------------
# DATE RANGE (if a specific time of year is desired)
# -----------------------------------------------------
start_date = datetime(YEAR,1,1)
end_date   = datetime(YEAR,12,31)
print(f"Processing period: {start_date.date()} → {end_date.date()}")

# -----------------------------
# COLLECTION L8+L9
# -----------------------------
l8 = ee.ImageCollection("LANDSAT/LC08/C02/T1_L2")
l9 = ee.ImageCollection("LANDSAT/LC09/C02/T1_L2")
l_all = l8.merge(l9)

col = (l_all
       .filterBounds(region)
       .filterDate(start_date, end_date)
       .filter(ee.Filter.lt("CLOUD_COVER", CLOUD_MAX))
       .filter(ee.Filter.eq("PROCESSING_LEVEL","L2SP"))
       .filter(ee.Filter.inList("SPACECRAFT_ID",["LANDSAT_8","LANDSAT_9"]))
)

count = col.size().getInfo()
print(f"\nFound {count} L2SP images.")

if count == 0:
    raise SystemExit("No images available.")

col_list = col.toList(count)

# -----------------------------------------------------------------------
# CALIBRATION FUNCTION FROM EXCEL in-situ measurements (AB–AC)
# -----------------------------------------------------------------------
def compute_calibration_from_excel(excel_path, x_col, y_col):
    """
    Read Excel columns by letter (AB / AC) using openpyxl and compute:
    y = a*x + b
    """
    print("\n=== CALIBRATION EXCEL ===")
    print("Loading:", excel_path)
    print("Using Excel columns:", x_col, y_col)

    wb = load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb.active

    x_vals = []
    y_vals = []

    for r in range(1, ws.max_row + 1):
        xv = ws[f"{x_col}{r}"].value
        yv = ws[f"{y_col}{r}"].value

        try:
            xv = float(xv)
            yv = float(yv)
        except:
            continue

        if np.isfinite(xv) and np.isfinite(yv):
            x_vals.append(xv)
            y_vals.append(yv)

    wb.close()

    x = np.asarray(x_vals, dtype="float64")
    y = np.asarray(y_vals, dtype="float64")

    if x.size < 2:
        raise ValueError("Not enough valid calibration points in Excel columns.")

    # Correct regression: in-situ (y) = a * satellite (x) + b
    A = np.vstack([x, np.ones_like(x)]).T
    a, b = np.linalg.lstsq(A, y, rcond=None)[0]

    y_pred = a * x + b
    ss_res = np.sum((y - y_pred)**2)
    ss_tot = np.sum((y - np.mean(y))**2)
    r2 = 1 - ss_res/ss_tot if ss_tot > 0 else np.nan
    r2_corr = np.corrcoef(x, y)[0, 1]**2

    print(f"Excel regression: y = {a:.4f} x + {b:.4f}")
    print(f"R² = {r2_corr:.4f}")
    print("Points used:", x.size)
    print("=============================\n")

    return a, b, r2

try:
    CAL_A, CAL_B, CAL_R2 = compute_calibration_from_excel(
        CALIB_EXCEL, CALIB_X_COL, CALIB_Y_COL
    )
except Exception as e:
    print("\n⚠ Calibration failed:", e)
    print("Using default regression: y = 1.1397 x - 0.7229")
    CAL_A, CAL_B, CAL_R2 = 1.1397, -0.7229, 0.9153

# -----------------------------
# USER CONFIRMATION
# -----------------------------
choice = input("Download and process these images? (y/n): ").lower()
if choice not in ["y","yes","o","oui"]:
    raise SystemExit("Cancelled")

# ----------------------------------
# Summary table (Landsat inventory)  
# ----------------------------------
records = []

for i in range(count):
    img = ee.Image(col_list.get(i))
    props = img.getInfo()["properties"]

    date_full = props.get("DATE_ACQUIRED", "N/A")
    time_utc  = props.get("SCENE_CENTER_TIME", "N/A")
    datetime_full = f"{date_full} {time_utc.replace('Z','')}"

    records.append({
        "Date & Time (UTC)": datetime_full,
        "Satellite": props.get("SPACECRAFT_ID", "N/A"),
        "Product ID": props.get("LANDSAT_PRODUCT_ID", "N/A"),
        "Cloud cover (%)": props.get("CLOUD_COVER", np.nan),
    })

df = pd.DataFrame(records)

print("\nLandsat scenes found:\n")
print(df.to_string(index=True))

excel_out = os.path.join(OUT_BASE, f"landsat_lugano_CIPAIS_inventory_{YEAR}.xlsx")
df.to_excel(excel_out, index=False)
print(f"\nExcel summary saved to:\n{excel_out}\n")

# -----------------------------
# PROCESS EACH IMAGE
# -----------------------------
for i in range(count):

    img = ee.Image(col_list.get(i))
    props = img.getInfo()["properties"]

    date = props["DATE_ACQUIRED"]
    sat  = props["SPACECRAFT_ID"]

    print(f"\n=== Image {i+1}/{count} ===")
    print(f"{date}  {sat}")
    print("============================")

    bands = img.bandNames().getInfo()

    # -----------------------------------------------------------
    # 1. ST_B10 BLOCK (optionally usable for final LST) 
    # -----------------------------------------------------------
    st_b10_available = False
    if "ST_B10" in bands:
        st_b10_available = True
        print("ST_B10 detected.")
        st = img.select("ST_B10")
        st_unscaled = st.multiply(0.00341802)

        try:
            kelvin_stats = st_unscaled.reduceRegion(
                reducer=ee.Reducer.minMax(),
                geometry=region,
                scale=30,
                maxPixels=1e9
            ).getInfo()

            print(f"ST_B10 diagnostic Kelvin range: min={kelvin_stats['ST_B10_min']:.1f}, max={kelvin_stats['ST_B10_max']:.1f}")

        except:
            print("ST_B10 diagnostic check failed.")
    else:
        print("ST_B10 missing: proceeding with Barsi TOA radiance method.")

    # -----------------------------------------
    # 2. QA + NIR MASK (FIXED: Fill + RADSAT)
    # -----------------------------------------
    qa = img.select("QA_PIXEL").toUint32()

    good = (
        qa.bitwiseAnd(1<<0).eq(0)   # Fill
        .And(qa.bitwiseAnd(1<<1).eq(0))  # Dilated cloud
        .And(qa.bitwiseAnd(1<<2).eq(0))  # Cirrus
        .And(qa.bitwiseAnd(1<<3).eq(0))  # Cloud
        .And(qa.bitwiseAnd(1<<4).eq(0))  # Cloud shadow
        .And(qa.bitwiseAnd(1<<5).eq(0))  # Snow
    )

    # Radiometric saturation (critical for thermal)
    good = good.And(img.select("QA_RADSAT").eq(0))

    nir = img.select("SR_B5")
    nir_ref = nir.multiply(0.0000275).add(-0.2)

    try:
        nir_stats = nir_ref.updateMask(good).reduceRegion(
            reducer=ee.Reducer.percentile([97.5]),
            geometry=lake_geom_ee,
            scale=30,
            maxPixels=1e9
        ).getInfo()

        key = [k for k in nir_stats if k.endswith("_p97_5")]
        if key and nir_stats[key[0]] is not None:
            nth = float(nir_stats[key[0]])
            print("NIR p97.5 =", nth)
            nir_mask = nir_ref.lt(nth)
        else:
            print("Invalid NIR percentile → disabling NIR mask")
            nir_mask = ee.Image(1)

    except:
        print("NIR percentile failed → disabling NIR mask")
        nir_mask = ee.Image(1)

    # ----------------------------------------------------------
    # 3. EXPORT TOA RADIANCE + ATMOSPHERIC BANDS (Barsi method)
    # ----------------------------------------------------------
    trad   = img.select("ST_TRAD")
    atrans = img.select("ST_ATRAN")
    drad   = img.select("ST_DRAD")
    urad   = img.select("ST_URAD")

    trad_s   = trad.multiply(0.001)
    atrans_s = atrans.multiply(ATRAN_SCALE)
    drad_s   = drad.multiply(DRAD_SCALE)
    urad_s   = urad.multiply(URAD_SCALE)

    atm_stack = trad_s.addBands(atrans_s).addBands(drad_s).addBands(urad_s)
    atm_masked = atm_stack.updateMask(good).updateMask(nir_mask).clip(lake_geom_ee)

    url = atm_masked.getDownloadURL({
    "scale": 30,
    "crs": "EPSG:4326",
    "region": gdf_ee.total_bounds.tolist(),
    "format": "GEO_TIFF",
    "filePerBand": False
    })

    rsp = requests.get(url, stream=True)

    # --- SAFETY CHECK: ensure a real GeoTIFF was returned ---
    if not rsp.ok or len(rsp.content) < 1000:
        print("Invalid download from Earth Engine.")
        print("HTTP status:", rsp.status_code)
        print("Response preview:", rsp.content[:300])
        continue

    tmp_tif = tempfile.NamedTemporaryFile(delete=False, suffix=".tif")
    tmp_tif.write(rsp.content)
    tmp_tif.close()

    # -----------------------------------------------------
    # 4. LOCAL LOADING OF TOA RADIANCE + ATMOSPHERIC TERMS
    # -----------------------------------------------------
    with rasterio.open(tmp_tif.name) as src:
        gdf_l = gdf_local if gdf_local.crs == src.crs else gdf_local.to_crs(src.crs)
        clipped, transform = mask(src, gdf_l.geometry, crop=True)
        meta = src.meta.copy()

    Lrad = clipped[0].astype("float32")
    TAU  = clipped[1].astype("float32")
    LD   = clipped[2].astype("float32")
    LU   = clipped[3].astype("float32")

        # --- SAFETY: remove invalid / fill-contaminated pixels ---
    invalid = (
        (Lrad <= 0) |
        (TAU  <= 0) |
        (LD   <= 0) |
        (LU   <= 0) |
        (~np.isfinite(Lrad)) |
        (~np.isfinite(TAU)) |
        (~np.isfinite(LD)) |
        (~np.isfinite(LU))
    )

    Lrad = np.where(invalid, np.nan, Lrad)
    TAU  = np.where(invalid, np.nan, TAU)
    LD   = np.where(invalid, np.nan, LD)
    LU   = np.where(invalid, np.nan, LU)

    # mean(scaled) == mean(DN)*scale (équivalent à l’Excel).
    TAU_s = np.nanmean(TAU)
    LD_s  = np.nanmean(LD)
    LU_s  = np.nanmean(LU)

    # -------------------------------------
    # 5. FULL BARSI ATMOSPHERIC CORRECTION
    # -------------------------------------
    print("Applying full Barsi atmospheric correction…")

    numerator = Lrad - LU_s - TAU_s * (1.0 - EMIS) * LD_s
    denom = TAU_s * EMIS

    denom = np.where((denom == 0) | (~np.isfinite(denom)), np.nan, denom)
    Ls = numerator / denom

    # >>> AJOUT MINIMAL: évite log invalide et scène “silencieuse” sans pixels physiques
    if not np.any(Ls > 0):
        print("No valid Ls > 0 after Barsi correction → skipping scene.")
        continue

    Tk = K1 / np.log((K2 / Ls) + 1.0)
    C = Tk - 273.15

    # ------------------------------------------------------
    # 5bis. OPTIONAL: use ST_B10 as final LST (if required)  
    # ------------------------------------------------------
    if USE_ST_B10_FOR_FINAL_LST and st_b10_available:
        print("Using ST_B10 for final LST (protocol option).")
        st_stack = st_unscaled.getDownloadURL({
            "scale": 30,
            "crs": "EPSG:4326",
            "region": gdf_ee.total_bounds.tolist(),
            "format": "GEO_TIFF",
            "filePerBand": False
        })
        rsp2 = requests.get(st_stack, stream=True)
        if not rsp2.ok or len(rsp2.content) < 1000:
            print("Invalid ST_B10 download → fallback to Barsi.")
        else:
            tmp_st = tempfile.NamedTemporaryFile(delete=False, suffix=".tif")
            tmp_st.write(rsp2.content)
            tmp_st.close()

            with rasterio.open(tmp_st.name) as src2:
                gdf_l2 = gdf_local if gdf_local.crs == src2.crs else gdf_local.to_crs(src2.crs)
                clipped2, transform2 = mask(src2, gdf_l2.geometry, crop=True)
            st_k = clipped2[0].astype("float32")
            st_c = st_k - 273.15
            C = np.where(np.isfinite(st_c), st_c, np.nan)

    # -----------------------------------------
    # 6. CALIBRATION AB–AC + 2.5–97.5 % FILTER
    # -----------------------------------------
    C_cal = CAL_A * C + CAL_B

    valid = C_cal[np.isfinite(C_cal)]
    if valid.size == 0:
        print("No valid calibrated LST pixels.")
        continue

    p2_5, p97_5 = np.nanpercentile(valid, [2.5, 97.5])
    C_clean = np.where((C_cal < p2_5) | (C_cal > p97_5), np.nan, C_cal)

    out_c = os.path.join(OUT_BASE, f"LST_Lugano_L2_clean_{sat}_{date.replace('-','')}.tif")
    meta.update(count=1, dtype="float32")   # CRS remains EPSG:4326 for now

    # -----------------------------------------
    # 7. REPROJECT FINAL OUTPUT TO EPSG:32632 
    # -----------------------------------------
    from rasterio.warp import calculate_default_transform, reproject, Resampling
    from rasterio.transform import array_bounds

    dst_crs = "EPSG:32632"

    # --- Compute bounds from transform (meta has no 'bounds') ---
    left, bottom, right, top = array_bounds(
        meta["height"],
        meta["width"],
        meta["transform"]
    )

    transform_utm, width_utm, height_utm = calculate_default_transform(
        meta["crs"],
        dst_crs,
        meta["width"],
        meta["height"],
        left, bottom, right, top
    )

    meta_utm = meta.copy()
    meta_utm.update({
        "crs": dst_crs,
        "transform": transform_utm,
        "width": width_utm,
        "height": height_utm,
        "count": 1,
        "dtype": "float32",
        "nodata": np.nan
    })

    with rasterio.open(out_c, "w", **meta_utm) as dst:
        reproject(
            source=C_clean.astype("float32"),
            destination=rasterio.band(dst, 1),
            src_transform=meta["transform"],
            src_crs=meta["crs"],
            dst_transform=transform_utm,
            dst_crs=dst_crs,
            resampling=Resampling.bilinear
        )

    print("Final °C raster:", out_c)
    time.sleep(PAUSE_S)

print("\n=== Completed with Barsi atmospheric LST correction ===")
